import psycopg2
from psycopg2 import Error
import openpyxl


try:
    # Подключение
    con = psycopg2.connect(user="postgres",
                                  password="1001",
                                  host="127.0.0.1",
                                  port="5432",
                                  database="encost_test")
    cursor = con.cursor()

    ### energy
    create_table_query = '''CREATE TABLE IF NOT EXISTS energy
                              (endpoint_id INT PRIMARY KEY NOT NULL,
                              event_time timestamptz); '''
    cursor.execute(create_table_query)
    con.commit()


    file = openpyxl.open('energy.csv', read_only=True)
    sheet = file.active

    for row in range(2, sheet.max_row + 1):
        endpoint_id = sheet[row][0].value
        event_time = sheet[row][1].value

        cursor.execute( "INSERT INTO energy (endpoint_id, event_time) VALUES (%s, %s) ON CONFLICT (endpoint_id) DO UPDATE SET endpoint_id = excluded.endpoint_id, event_time = excluded.event_time;",(endpoint_id, event_time))
        con.commit()


    ### operators
    create_table_query = '''CREATE TABLE IF NOT EXISTS operators
                                  (endpoint_id INT PRIMARY KEY NOT NULL,
                                  login_time timestamptz,
                                  logout_time timestamptz,
                                  operator_name, text); '''
    cursor.execute(create_table_query)
    con.commit()

    file = openpyxl.open('operators.csv', read_only=True)
    sheet = file.active

    for row in range(2, sheet.max_row + 1):
        cursor.execute(
            "INSERT INTO operators (endpoint_id, login_time, logout_time, operator_name ) VALUES (%s, %s, %s, %s)",(sheet[row][0].value, sheet[row][1].value, sheet[row][2].value, sheet[row][3].value))
        con.commit()

    ###### periods
    create_table_query = '''CREATE TABLE IF NOT EXISTS periods
                                      (endpoint_id INT PRIMARY KEY NOT NULL,
                                      mode_start timestamptz,
                                      mode_duration integer,
                                      label, text); '''
    cursor.execute(create_table_query)
    con.commit()

    file = openpyxl.open('periods.csv', read_only=True)
    sheet = file.active

    for row in range(2, sheet.max_row + 1):
        cursor.execute(
            "INSERT INTO periods (endpoint_id, mode_start, mode_duration, label ) VALUES (%s, %s, %s, %s)",
            (sheet[row][0].value, sheet[row][1].value, sheet[row][2].value, sheet[row][3].value))
        con.commit()


    ### reasons
    create_table_query = '''CREATE TABLE IF NOT EXISTS reasons
                                          (endpoint_id INT PRIMARY KEY NOT NULL,
                                          event_time timestamptz,
                                          
                                          reason, text); '''
    cursor.execute(create_table_query)
    con.commit()

    file = openpyxl.open('reasons.csv', read_only=True)
    sheet = file.active

    for row in range(2, sheet.max_row + 1):
        cursor.execute(
            "INSERT INTO reasons (endpoint_id, event_time, reason ) VALUES (%s, %s, %s)",
            (sheet[row][0].value, sheet[row][1].value, sheet[row][2].value))
        con.commit()


    ### Запрос на представление
    cursor.execute("CREATE VIEW period_view.sql AS SELECT endpoint_id, mode_start, mode_end, mode_duration, label, reason, operator_name"
                   "energy_sum FROM reasons, energy, operators, periods WHERE reasons.endpoint_id =energy.endpoint_id AND energy.endpoint_id = operators.endpoint_id"
                   " AND operators.endpoint_id = periods.endpoint_id  GROUP BY endpoint_id" )

except (Exception, Error) as error:
    print("Ошибка при работе с PostgreSQL", error)
finally:
    if con:
        cursor.close()
        con.close()





